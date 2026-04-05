"""
Build pre-game normalization rules from the shared dictionary and known AI patterns.

Professional game localization involves three categories of normalization rules
that can be prepared BEFORE the game is even available:

1. TERMINOLOGY CONSISTENCY - Same Chinese term should always translate the same way.
   Mined from our shared dictionary where the same Chinese text had conflicting
   translations across WOJD namespaces (we picked the majority, but the minority
   translations will likely resurface when AI translates the new game).

2. AI MISTAKE PATTERNS - Known systematic errors that AI translators make when
   translating Chinese gaming terminology. These are predictable and repeatable.

3. STYLE GUIDE RULES - Consistent formatting, capitalization, and phrasing standards
   for English game text (e.g. "Lv" vs "Level", singular vs plural for UI labels).

Output: translations/normalization_rules.xlsx
Format: Simp Chinese | Trad Chinese | Good Translation | Bad Translation | Category | Notes

Usage:
    python build_normalization_rules.py [--dictionary PATH] [--output PATH]
"""

import json
import os
import sys
import argparse
from collections import defaultdict, Counter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from lib.utils import contains_chinese, count_chinese_chars

DEFAULT_WOJD_MAP = "D:/Games/WOJD/EngPatch/PATCHER/chinese_english_map.json"
DEFAULT_OUTPUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                               "translations", "normalization_rules.xlsx")


# ============================================================================
# CATEGORY 1: Terminology Consistency
# Mine the WOJD map for Chinese terms with multiple conflicting translations.
# These conflicts WILL reappear in the new game when AI translates them again.
# ============================================================================

def mine_terminology_conflicts(wojd_map_path, min_occurrences=3):
    """
    Find Chinese terms that had conflicting English translations in WOJD.

    These are terms where the inconsistency detector (Stage 6) would have
    flagged them. By pre-loading these rules, we catch the conflicts before
    they happen in the new game.

    Args:
        wojd_map_path: Path to WOJD's chinese_english_map.json
        min_occurrences: Minimum times the "bad" translation appeared (filters noise)

    Returns:
        List of rule dicts with Chinese, Good, Bad, counts.
    """
    print(f"Mining terminology conflicts from WOJD map...")

    with open(wojd_map_path, 'r', encoding='utf-8-sig') as f:
        wojd_map = json.load(f)

    # Collect all translations for each Chinese string
    term_translations = defaultdict(list)
    for namespace, entries in wojd_map.items():
        if not isinstance(entries, dict):
            continue
        for chinese, english in entries.items():
            if not isinstance(english, str) or not english.strip():
                continue
            chinese = chinese.strip()
            english = english.strip()
            if not contains_chinese(chinese):
                continue
            # Focus on short-to-medium terms (UI labels, skill names, item names)
            cn_chars = count_chinese_chars(chinese)
            if cn_chars < 1 or cn_chars > 30:
                continue
            term_translations[chinese].append(english)

    # Find terms with conflicting translations
    rules = []
    for chinese, translations in term_translations.items():
        counts = Counter(translations)
        if len(counts) < 2:
            continue  # Only one translation, no conflict

        # Sort by frequency
        sorted_translations = counts.most_common()
        good_translation = sorted_translations[0][0]
        good_count = sorted_translations[0][1]

        for bad_translation, bad_count in sorted_translations[1:]:
            if bad_count < min_occurrences:
                continue
            if good_translation.lower() == bad_translation.lower():
                continue  # Just case difference, skip
            rules.append({
                'chinese': chinese,
                'good': good_translation,
                'bad': bad_translation,
                'good_count': good_count,
                'bad_count': bad_count,
                'category': 'Terminology',
            })

    print(f"  Found {len(rules)} terminology conflict rules")
    return rules


# ============================================================================
# CATEGORY 2: AI Mistake Patterns
# Known systematic errors AI translators make with Chinese gaming terms.
# These are universal across games, not WOJD-specific.
# ============================================================================

def get_ai_mistake_rules():
    """
    Return known AI translation mistake patterns for Chinese game text.

    These are mistakes observed across multiple games and AI models.
    They apply universally to Chinese->English game localization.
    """
    rules = []

    # --- Common gaming term consistency ---
    gaming_terms = [
        # (Chinese, Good, Bad, Notes)
        ("等级", "Level", "Grade", "RPG standard: 等级 is always Level"),
        ("等级", "Level", "Rank", "Rank is for competitive/PvP, not character level"),
        ("经验", "EXP", "Experience", "UI brevity: use EXP in game UI"),
        ("经验值", "EXP", "Experience Points", "UI brevity"),
        ("生命值", "HP", "Health Points", "UI brevity"),
        ("生命值", "HP", "Life Points", "UI brevity"),
        ("生命", "HP", "Life", "In combat context, 生命 means HP not Life"),
        ("魔法值", "MP", "Magic Points", "UI brevity"),
        ("攻击力", "ATK", "Attack Power", "UI brevity for stat labels"),
        ("防御力", "DEF", "Defense Power", "UI brevity for stat labels"),
        ("暴击", "CRIT", "Critical Hit", "UI brevity"),
        ("暴击率", "CRIT Rate", "Critical Hit Rate", "UI brevity"),
        ("暴击伤害", "CRIT DMG", "Critical Hit Damage", "UI brevity"),
        ("冷却", "Cooldown", "Cool Down", "One word, not two"),
        ("冷却时间", "Cooldown", "Cooling Time", "Gaming standard"),
        ("冷却时间", "Cooldown", "Cool-down Time", "Gaming standard"),
        ("副本", "Dungeon", "Instance", "Western gaming term preference"),
        ("副本", "Dungeon", "Copy", "Literal translation, incorrect"),
        ("坐骑", "Mount", "Riding Pet", "Western gaming standard"),
        ("坐骑", "Mount", "Steed", "Mount is the universal gaming term"),
        ("背包", "Inventory", "Backpack", "Gaming UI standard"),
        ("背包", "Inventory", "Bag", "Gaming UI standard for full inventory"),
        ("任务", "Quest", "Task", "RPG standard: 任务 in story context = Quest"),
        ("日常任务", "Daily Quest", "Daily Task", "RPG standard"),
        ("成就", "Achievements", "Achievement", "Plural for the system/menu label"),
        ("装备", "Equipment", "Gear", "Equipment for the category, Gear is informal"),
        ("强化", "Enhance", "Strengthen", "Gaming standard for gear upgrade"),
        ("升级", "Upgrade", "Level Up", "Context: Upgrade for items, Level Up for characters"),
        ("充值", "Top Up", "Recharge", "Standard for adding premium currency"),
        ("钻石", "Diamonds", "Diamond", "Plural for currency"),
        ("金币", "Gold", "Gold Coins", "UI brevity"),
        ("银币", "Silver", "Silver Coins", "UI brevity"),
        ("体力", "Stamina", "Physical Strength", "Gaming standard for energy system"),
        ("体力", "Stamina", "Physical Power", "Gaming standard"),
        ("好友", "Friends", "Good Friends", "Literal translation, incorrect for UI"),
        ("公会", "Guild", "Union", "Western gaming standard"),
        ("工会", "Guild", "Union", "Western gaming standard"),
        ("排行榜", "Leaderboard", "Ranking", "Western UI standard"),
        ("排行榜", "Leaderboard", "Rankings", "Singular for the feature name"),
        ("商城", "Shop", "Mall", "Gaming standard for in-game store"),
        ("商店", "Shop", "Store", "Shop is more common in games"),
        ("抽奖", "Gacha", "Lottery", "Gacha is the established term"),
        ("抽奖", "Gacha", "Draw", "Too vague"),
        ("抽卡", "Gacha", "Card Draw", "Gacha is the established term"),
        ("首充", "First Purchase", "First Recharge", "Standard monetization term"),
        ("签到", "Check-in", "Sign In", "Daily check-in, not account sign-in"),
        ("签到", "Check-in", "Sign-in", "Daily check-in, not account sign-in"),
        ("邮件", "Mail", "Email", "In-game mail, not email"),
        ("设置", "Settings", "Setup", "UI standard"),
        ("确认", "Confirm", "Determine", "AI often mistranslates 确认"),
        ("取消", "Cancel", "Cancellation", "UI button text"),
        ("返回", "Back", "Return", "UI navigation standard"),
        ("关闭", "Close", "Shut Down", "UI standard"),
        ("领取", "Claim", "Receive", "Gaming standard for rewards"),
        ("领取", "Claim", "Collect", "Claim is more standard for reward buttons"),
    ]

    for chinese, good, bad, notes in gaming_terms:
        rules.append({
            'chinese': chinese,
            'good': good,
            'bad': bad,
            'good_count': 0,
            'bad_count': 0,
            'category': 'AI Pattern',
            'notes': notes,
        })

    # --- Plural/singular consistency for UI labels ---
    # In game UI, menu/section labels are typically plural, button actions are singular
    plural_rules = [
        ("技能", "Skills", "Skill", "Menu label should be plural"),
        ("道具", "Items", "Item", "Menu label should be plural"),
        ("宠物", "Pets", "Pet", "Menu label should be plural"),
        ("称号", "Titles", "Title", "Menu label should be plural"),
        ("时装", "Outfits", "Outfit", "Menu label should be plural"),
        ("翅膀", "Wings", "Wing", "Menu label should be plural"),
        ("坐骑", "Mounts", "Mount", "Menu label should be plural"),
    ]

    for chinese, good, bad, notes in plural_rules:
        rules.append({
            'chinese': chinese,
            'good': good,
            'bad': bad,
            'good_count': 0,
            'bad_count': 0,
            'category': 'Style Guide',
            'notes': notes + " (context-dependent: plural for menus, singular for tooltips)",
        })

    print(f"  Built {len(rules)} AI mistake pattern rules")
    return rules


# ============================================================================
# CATEGORY 3: Style Guide Rules
# Formatting and phrasing standards for English game text.
# ============================================================================

def get_style_guide_rules():
    """
    Return style guide enforcement rules.

    These enforce consistent formatting across all translations.
    """
    rules = []

    formatting_rules = [
        # Level formatting
        ("{0}级", "Lv.{0}", "Lv {0}", "Style Guide", "Consistent level abbreviation with period"),
        ("{0}级", "Lv.{0}", "Level {0}", "Style Guide", "UI brevity: Lv. not Level in compact UI"),
        ("{0}级", "Lv.{0}", "LV{0}", "Style Guide", "Consistent casing: Lv. not LV"),
        ("{0}级", "Lv.{0}", "lv.{0}", "Style Guide", "Capitalize: Lv. not lv."),

        # Percentage formatting
        ("{0}%", "{0}%", "{0} %", "Style Guide", "No space before percent sign"),

        # Time formatting
        ("{0}秒", "{0}s", "{0} sec", "Style Guide", "Brevity: s not sec"),
        ("{0}秒", "{0}s", "{0} seconds", "Style Guide", "Brevity: s not seconds in UI"),
        ("{0}分钟", "{0}min", "{0} minutes", "Style Guide", "Brevity in compact UI"),

        # Number formatting
        ("{0}次", "{0}x", "{0} times", "Style Guide", "Brevity: x not times"),
        ("{0}个", "{0}x", "{0} pieces", "Style Guide", "Brevity for generic counter"),
    ]

    for chinese, good, bad, category, notes in formatting_rules:
        rules.append({
            'chinese': chinese,
            'good': good,
            'bad': bad,
            'good_count': 0,
            'bad_count': 0,
            'category': category,
            'notes': notes,
        })

    print(f"  Built {len(rules)} style guide rules")
    return rules


# ============================================================================
# Output
# ============================================================================

def save_rules_xlsx(rules, output_path):
    """Save rules as an Excel spreadsheet for review and editing."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
        # Fall back to CSV
        save_rules_csv(rules, output_path.replace('.xlsx', '.csv'))
        return

    try:
        import opencc
        s2t = opencc.OpenCC('s2t')
        has_opencc = True
    except ImportError:
        has_opencc = False
        print("WARNING: opencc not installed. Traditional Chinese column will be empty.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "normalization_rules"

    # Header row
    headers = ['Simp Chinese', 'Trad Chinese', 'Good Translation', 'Bad Translation']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Data rows
    for row_idx, rule in enumerate(rules, 2):
        chinese = rule['chinese']
        trad = s2t.convert(chinese) if has_opencc else ''

        ws.cell(row=row_idx, column=1, value=chinese)
        ws.cell(row=row_idx, column=2, value=trad)
        ws.cell(row=row_idx, column=3, value=rule['good'])
        ws.cell(row=row_idx, column=4, value=rule['bad'])

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"\nSaved {len(rules)} rules to: {output_path}")


def save_rules_csv(rules, output_path):
    """Fallback: save as CSV if openpyxl not available."""
    import csv

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Simp Chinese', 'Trad Chinese', 'Good Translation',
                         'Bad Translation'])
        for rule in rules:
            writer.writerow([
                rule['chinese'], '', rule['good'], rule['bad']
            ])
    print(f"\nSaved {len(rules)} rules to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Build pre-game normalization rules")
    parser.add_argument("--wojd-map", default=DEFAULT_WOJD_MAP,
                        help="Path to WOJD chinese_english_map.json")
    parser.add_argument("--output", default=DEFAULT_OUTPUT,
                        help="Output path for normalization_rules.xlsx")
    parser.add_argument("--min-occurrences", type=int, default=3,
                        help="Minimum bad translation occurrences to include (default: 3)")
    parser.add_argument("--skip-wojd", action="store_true",
                        help="Skip WOJD terminology mining (only include AI patterns + style guide)")
    args = parser.parse_args()

    print("=== Building Normalization Rules ===\n")

    all_rules = []

    # Category 1: Terminology from WOJD
    if not args.skip_wojd and os.path.exists(args.wojd_map):
        terminology_rules = mine_terminology_conflicts(args.wojd_map, args.min_occurrences)
        all_rules.extend(terminology_rules)
    elif not args.skip_wojd:
        print(f"  WOJD map not found at: {args.wojd_map} (skipping terminology mining)")

    # Category 2: AI mistake patterns
    ai_rules = get_ai_mistake_rules()
    all_rules.extend(ai_rules)

    # Category 3: Style guide
    style_rules = get_style_guide_rules()
    all_rules.extend(style_rules)

    # Build lookup of curated rules by their good/bad translations per Chinese term.
    # Include S/T variants so Traditional-Chinese Terminology rules are also caught.
    try:
        import opencc
        s2t = opencc.OpenCC('s2t')
        t2s = opencc.OpenCC('t2s')
        has_opencc = True
    except ImportError:
        has_opencc = False

    curated_good_bad = defaultdict(set)  # chinese -> set of (good, bad) tuples
    for rule in all_rules:
        if rule.get('category') in ('Style Guide', 'AI Pattern'):
            curated_good_bad[rule['chinese']].add((rule['good'], rule['bad']))
            if has_opencc:
                # Also register for S/T variants
                curated_good_bad[s2t.convert(rule['chinese'])].add((rule['good'], rule['bad']))
                curated_good_bad[t2s.convert(rule['chinese'])].add((rule['good'], rule['bad']))

    # Filter out Terminology rules that contradict curated rules
    # A contradiction: Terminology says "X is good, Y is bad" but a curated rule
    # for the same term (or its S/T variant) says "Y is good, X is bad"
    filtered = []
    dropped = 0
    for rule in all_rules:
        if rule.get('category') == 'Terminology':
            curated_set = curated_good_bad.get(rule['chinese'], set())
            dominated = False
            for curated_good, curated_bad in curated_set:
                if rule['good'] == curated_bad or rule['bad'] == curated_good:
                    dominated = True
                    break
            if dominated:
                dropped += 1
                continue
        filtered.append(rule)

    if dropped:
        print(f"\n  Dropped {dropped} Terminology rules that contradicted Style Guide/AI Pattern rules")

    # Deduplicate remaining (same Chinese + same Bad = keep the one with higher priority)
    seen = {}
    deduped = []
    for rule in filtered:
        key = (rule['chinese'], rule['bad'])
        if key in seen:
            existing = seen[key]
            # Prefer curated categories over Terminology
            curated_cats = ('Style Guide', 'AI Pattern')
            if rule.get('category') in curated_cats and existing.get('category') not in curated_cats:
                deduped.remove(existing)
                seen[key] = rule
                deduped.append(rule)
            elif rule['bad_count'] > existing['bad_count'] and rule.get('category') == existing.get('category'):
                deduped.remove(existing)
                seen[key] = rule
                deduped.append(rule)
        else:
            seen[key] = rule
            deduped.append(rule)

    print(f"\n  Total: {len(deduped)} rules after deduplication")

    # Sort: Style Guide first, then AI Patterns, then Terminology (by bad_count desc)
    category_order = {'Style Guide': 0, 'AI Pattern': 1, 'Terminology': 2}
    deduped.sort(key=lambda r: (category_order.get(r.get('category', ''), 99), -r['bad_count']))

    save_rules_xlsx(deduped, args.output)

    # Print summary
    by_category = defaultdict(int)
    for rule in deduped:
        by_category[rule.get('category', 'Unknown')] += 1
    print("\n--- Rules by Category ---")
    for cat, count in sorted(by_category.items()):
        print(f"  {cat}: {count}")


if __name__ == "__main__":
    main()
